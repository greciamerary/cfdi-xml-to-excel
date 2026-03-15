import xml.etree.ElementTree as ET
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk, ImageSequence
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def seleccionar_archivo(tipo, extension):
    """Abre un cuadro de diálogo para seleccionar un archivo"""
    root_tk = tk.Tk()
    root_tk.withdraw()
    archivo = filedialog.askopenfilename(title=f"Hola RH, selecciona un archivo {tipo}", filetypes=[(f"Archivos {tipo}", extension)])
    return archivo

# Seleccionar archivos XML y Excel
archivo_xml = seleccionar_archivo("XML", "*.xml")
if not archivo_xml:
    print("No se seleccionó ningún archivo XML.")
    exit()

archivo_excel = seleccionar_archivo("Excel", "*.xlsx")
if not archivo_excel:
    print("No se seleccionó ningún archivo Excel.")
    exit()

try:
    # Cargar el XML
    tree = ET.parse(archivo_xml)
    root = tree.getroot()
    
    # Espacio de nombres
    ns = {'cfdi': 'http://www.sat.gob.mx/cfd/3'}

    # Extraer datos
    folio = root.attrib.get("Folio", "")
    fecha = root.attrib.get("Fecha", "")

    # Nombre del proveedor (Emisor)
    emisor = root.find("cfdi:Emisor", ns)
    nombre_proveedor = emisor.attrib.get("Nombre", "") if emisor is not None else ""

    # Recorrer los productos dentro de <cfdi:Conceptos>
    datos = []
    for concepto in root.findall(".//cfdi:Concepto", ns):
        fila = {
            "Folio": folio,
            "Fecha": fecha,
            "Nombre (proveedor)": nombre_proveedor,
            "No Identificación": concepto.attrib.get("NoIdentificacion", ""),
            "Valor unitario": concepto.attrib.get("ValorUnitario", ""),
        }
        datos.append(fila)

    # Convertir a DataFrame
    df_nuevo = pd.DataFrame(datos)

    # Cargar el archivo Excel existente y agregar los nuevos datos
    if os.path.exists(archivo_excel):
        df_existente = pd.read_excel(archivo_excel, engine="openpyxl")
        df_final = pd.concat([df_existente, df_nuevo], ignore_index=True)
    else:
        df_final = df_nuevo

    # Guardar el Excel actualizado
    df_final.to_excel(archivo_excel, index=False, engine="openpyxl")

    # Ajustar el ancho de las columnas en el Excel
    wb = load_workbook(archivo_excel)
    ws = wb.active

    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2  # Añadir un poco de espacio extra

    wb.save(archivo_excel)

    # Mostrar ventana emergente con GIF animado
    def mostrar_popup():
        popup = tk.Toplevel()
        popup.title("Proceso Completado")
        popup.geometry("300x200")

        label_mensaje = tk.Label(popup, text="Datos agregados con éxito!", font=("Arial", 12))
        label_mensaje.pack(pady=10)

        try:
            gif = Image.open("animacion.gif")  # Asegúrate de tener un archivo llamado 'animacion.gif'

            # Extraer fotogramas del GIF
            frames = [ImageTk.PhotoImage(frame.copy()) for frame in ImageSequence.Iterator(gif)]
            
            label_gif = tk.Label(popup)
            label_gif.pack()

            # Función para actualizar el GIF
            def actualizar_gif(ind=0):
                frame = frames[ind]
                label_gif.config(image=frame)
                popup.after(100, actualizar_gif, (ind + 1) % len(frames))  # Cambia cada 100ms

            actualizar_gif()  # Iniciar animación

        except Exception as e:
            messagebox.showwarning("Advertencia", f"No se pudo cargar la imagen: {e}")

        cerrar_btn = tk.Button(popup, text="Cerrar", command=popup.destroy)
        cerrar_btn.pack(pady=10)

    # Ventana principal visible
    root = tk.Tk()
    root.withdraw()  # Ocultar la ventana principal

    # Llamar a la ventana emergente
    mostrar_popup()

    # Iniciar el bucle de eventos de Tkinter para que el popup se muestre correctamente
    root.mainloop()

    print(f"Datos agregados y columnas ajustadas en {archivo_excel}.")

except FileNotFoundError:
    print("Error: No se encontró el archivo.")
except Exception as e:
    print(f"Ocurrió un error: {e}")
